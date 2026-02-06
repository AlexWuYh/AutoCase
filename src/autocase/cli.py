import argparse
import csv
import json
import os
import sys
import time
from typing import Optional, List, Iterable

from .parser import parse_casespecs_yaml, load_yaml
from .generator import to_excel_rows, TestCase, llm_items_to_cases, cases_to_json
from .llm_client import generate_llm_cases, generate_module_code

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
except Exception:  # pragma: no cover - optional import guard
    openpyxl = None


def _read_input(path: Optional[str]) -> str:
    if path:
        with open(path, "r", encoding="utf-8") as f:
            return f.read()
    return sys.stdin.read()


def _collect_input_files(raw_items: Optional[Iterable[str]]) -> List[str]:
    if not raw_items:
        return []
    input_dir = "inputs"
    collected: List[str] = []
    for raw in raw_items:
        if not raw:
            continue
        path = raw
        if not os.path.isabs(path) and not os.path.exists(path):
            candidate = os.path.join(input_dir, path)
            if os.path.exists(candidate):
                path = candidate
        if os.path.isdir(path):
            for name in sorted(os.listdir(path)):
                if name.lower().endswith((".yaml", ".yml")):
                    collected.append(os.path.join(path, name))
        else:
            collected.append(path)
    seen = set()
    unique: List[str] = []
    for p in collected:
        if p not in seen:
            seen.add(p)
            unique.append(p)
    return unique

def _print_banner() -> None:
    banner = r"""
    _         _              _
   / \\  _   _| |_ ___   ___ | |___  ___
  / _ \\| | | | __/ _ \\ / _ \\| / __|/ _ \\
 / ___ \\ |_| | || (_) | (_) | \\__ \\  __/
/_/   \\_\\__,_|\\__\\___/ \\___/|_|___/\\___|
    """
    meta = "Author: Yinghao Wu | GitHub: https://github.com/AlexWuYh/AutoCase"
    if sys.stdout.isatty():
        blue = "\033[94m"
        reset = "\033[0m"
        print(blue + banner + reset)
        print(meta)
    else:
        print(banner)
        print(meta)


def _supports_color() -> bool:
    return sys.stderr.isatty() and os.getenv("NO_COLOR") is None


def _color(text: str, code: str) -> str:
    if not _supports_color():
        return text
    return f"\033[{code}m{text}\033[0m"


def _log_header(title: str) -> None:
    line = "─" * 60
    print(_color(f"╭{line}╮", "38;5;33"), file=sys.stderr)
    print(_color(f"│ {title:<58} │", "1;38;5;33"), file=sys.stderr)
    print(_color(f"╰{line}╯", "38;5;33"), file=sys.stderr)


def _log_kv(key: str, value: str) -> None:
    k = _color(f"{key:<10}", "1;38;5;39")
    v = _color(value, "38;5;252")
    print(f"  {k} {v}", file=sys.stderr)


def _log_step(prefix: str, message: str) -> None:
    p = _color(prefix, "1;38;5;40")
    m = _color(message, "38;5;252")
    print(f"  {p} {m}", file=sys.stderr)


def _log_level(level: str, message: str) -> None:
    color_map = {"INFO": "38;5;39", "WARN": "38;5;214", "ERROR": "38;5;196"}
    lvl = _color(f"[{level}]", f"1;{color_map.get(level, '38;5;39')}")
    msg = _color(message, "38;5;252")
    print(f"  {lvl} {msg}", file=sys.stderr)


def _progress_bar(current: int, total: int, width: int = 24) -> str:
    if total <= 0:
        total = 1
    ratio = min(max(current / total, 0), 1)
    filled = int(ratio * width)
    bar = "█" * filled + "░" * (width - filled)
    return f"{bar} {int(ratio * 100):>3d}%"


def _progress_update(line: str, done: bool = False) -> None:
    if not _supports_color():
        return
    end = "\n" if done else ""
    sys.stderr.write("\r" + line + (" " * 4) + end)
    sys.stderr.flush()


def _load_module_cache(path: str) -> dict:
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
            return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _save_module_cache(path: str, data: dict) -> None:
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def main() -> int:
    parser = argparse.ArgumentParser(description="AutoCase - 生成标准测试用例表格(Excel/CSV)")
    parser.add_argument(
        "-f",
        "--file",
        action="append",
        help="输入YAML文件路径（可重复）；也可传目录以加载该目录下所有 .yaml/.yml",
    )
    parser.add_argument(
        "--llm-config",
        default="config/llm.yaml",
        help="大模型参数配置文件路径（当前版本仅读取校验）",
    )
    parser.add_argument(
        "--prompt",
        default="config/system_prompt.txt",
        help="系统级prompt文件路径（可编辑）",
    )
    parser.add_argument(
        "-o",
        "--output",
        default="output.xlsx",
        help="输出文件路径（支持 .xlsx 或 .csv，默认 .xlsx）",
    )
    parser.add_argument(
        "--no-banner",
        action="store_true",
        help="关闭启动Banner显示",
    )
    parser.add_argument(
        "--json-only",
        action="store_true",
        help="仅输出JSON到STDOUT，不生成Excel",
    )
    args = parser.parse_args()

    if len(sys.argv) == 1:
        _print_banner()
        parser.print_help()
        print("Docs: https://github.com/AlexWuYh/AutoCase")
        return 0

    if not args.no_banner:
        _print_banner()

    input_paths = _collect_input_files(args.file)
    if not input_paths and sys.stdin.isatty():
        parser.print_help()
        return 0

    start_ts = time.perf_counter()
    _log_header("AutoCase Run")
    _log_kv("Status", "start  🚀")
    if input_paths:
        _log_kv("Inputs", f"{len(input_paths)} file(s)")
    else:
        _log_kv("Inputs", "STDIN")

    output_dir = "outputs"
    for p in input_paths:
        if not os.path.exists(p):
            print(f"输入文件不存在: {p}", file=sys.stderr)
            return 2

    output_path = args.output
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    if output_path == "output.xlsx" and input_paths:
        if len(input_paths) == 1:
            base = os.path.splitext(os.path.basename(input_paths[0]))[0]
            output_path = os.path.join(output_dir, f"{base}_{timestamp}_testcases.xlsx")
        else:
            output_path = os.path.join(output_dir, f"combined_{timestamp}_testcases.xlsx")
    elif output_path == "output.xlsx" and not input_paths:
        output_path = os.path.join(output_dir, f"stdin_{timestamp}_testcases.xlsx")
    elif not os.path.isabs(output_path):
        output_path = os.path.join(output_dir, output_path)

    _log_header("Inputs")
    if input_paths:
        _log_kv("Count", str(len(input_paths)))
        for p in input_paths:
            _log_step("•", p)
    else:
        _log_kv("Count", "STDIN")

    # Read LLM config & system prompt (for validation/visibility)
    _log_kv("Config", args.llm_config)
    if not os.path.exists(args.llm_config):
        print(f"LLM 配置文件不存在: {args.llm_config}", file=sys.stderr)
        return 2
    llm_text = _read_input(args.llm_config)
    try:
        llm_config = load_yaml(llm_text)
    except ValueError as e:
        print(f"LLM 配置错误: {e}", file=sys.stderr)
        return 2

    if not bool(llm_config.get("enabled", True)):
        print("LLM 已禁用，请在 llm.yaml 中设置 enabled: true", file=sys.stderr)
        return 2

    _log_kv("Prompt", args.prompt)
    if not os.path.exists(args.prompt):
        print(f"系统级prompt文件不存在: {args.prompt}", file=sys.stderr)
        return 2
    prompt_text = _read_input(args.prompt)
    if not prompt_text.strip():
        print("系统级prompt为空", file=sys.stderr)
        return 2

    all_specs = []
    if not input_paths:
        text = _read_input(None)
        if not text.strip():
            print("未读取到输入内容", file=sys.stderr)
            return 2
        try:
            _log_step("✓", "Parse YAML")
            specs = parse_casespecs_yaml(text)
        except ValueError as e:
            print(str(e), file=sys.stderr)
            return 2
        all_specs.extend(specs)
    else:
        for p in input_paths:
            text = _read_input(p)
            if not text.strip():
                print(f"未读取到输入内容: {p}", file=sys.stderr)
                return 2
            try:
                _log_step("✓", f"Parse YAML: {os.path.basename(p)}")
                specs = parse_casespecs_yaml(text)
            except ValueError as e:
                print(f"{p} 解析失败: {e}", file=sys.stderr)
                return 2
            for spec in specs:
                spec.source = os.path.basename(p)
            all_specs.extend(specs)

    module_code_cache = {}
    cache_path = ".autocase_module_cache.json"
    module_code_cache.update(_load_module_cache(cache_path))
    for spec in all_specs:
        if not spec.module_code:
            if spec.module not in module_code_cache:
                try:
                    module_code_cache[spec.module] = generate_module_code(
                        spec.module, llm_config
                    )
                    _log_level(
                        "INFO",
                        f"模块缩写生成: {spec.module} -> {module_code_cache[spec.module]}",
                    )
                except Exception as e:
                    _log_level(
                        "WARN",
                        f"模块缩写生成失败: {spec.module} ({e})，使用 MOD",
                    )
                    module_code_cache[spec.module] = "MOD"
            spec.module_code = module_code_cache[spec.module]
    _save_module_cache(cache_path, module_code_cache)

    cases: List[TestCase] = []
    next_index = 1
    total = len(all_specs)
    _log_header("Generation")
    for idx, spec in enumerate(all_specs, start=1):
        step_start = time.perf_counter()
        bar = _progress_bar(idx - 1, total)
        display_src = spec.source or "STDIN"
        display_title = f"{display_src} | {spec.feature}"
        if _supports_color():
            line = _color("▸", "1;38;5;40") + " " + _color(
                f"{bar}  {idx}/{total} | {display_title}  (start)",
                "38;5;252",
            )
            _progress_update(line)
        else:
            _log_step("▸", f"{bar}  {idx}/{total} | {display_title}  (start)")
        try:
            llm_items = generate_llm_cases(spec, llm_config, prompt_text)
            llm_cases, next_index = llm_items_to_cases(llm_items, spec, next_index)
            cases.extend(llm_cases)
        except Exception as e:
            _log_level("ERROR", f"LLM 生成失败: {e}")
            return 2
        step_elapsed = time.perf_counter() - step_start
        bar_done = _progress_bar(idx, total)
        if _supports_color():
            line = _color("✓", "1;38;5;40") + " " + _color(
                f"{bar_done}  {idx}/{total} | {display_title}  ({step_elapsed:.2f}s)",
                "38;5;252",
            )
            _progress_update(line, done=True)
        else:
            _log_step("✓", f"{bar_done}  {idx}/{total} | {display_title}  ({step_elapsed:.2f}s)")
    if args.json_only:
        elapsed = time.perf_counter() - start_ts
        _log_header("Result")
        _log_kv("Cases", f"{len(cases)} (JSON)")
        _log_kv("Output", "JSON to STDOUT (no Excel)")
        _log_kv("Elapsed", f"{elapsed:.2f}s  ⏱️")
        _log_kv("Status", "done  ✅")
        print(json.dumps(cases_to_json(cases), ensure_ascii=False, indent=2))
        return 0

    output_parent = os.path.dirname(output_path)
    if output_parent:
        os.makedirs(output_parent, exist_ok=True)

    _log_header("Output")
    _log_kv("Write", output_path)
    rows = to_excel_rows(cases)
    output_ext = os.path.splitext(output_path)[1].lower()
    if output_ext == ".csv":
        with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerows(rows)
        print(f"已生成: {output_path}")
        return 0

    if openpyxl is None:
        print("缺少依赖: openpyxl，请先安装依赖", file=sys.stderr)
        return 2
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TestCases"
    for row in rows:
        ws.append(row)

    # Styling
    header_fill = PatternFill("solid", fgColor="E8EEF7")
    header_font = Font(bold=True)
    thin = Side(border_style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = wrap
            cell.border = border
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
            elif cell.row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8FAFC")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(output_path)
    print(f"已生成: {output_path}")
    elapsed = time.perf_counter() - start_ts
    _log_kv("Elapsed", f"{elapsed:.2f}s  ⏱️")
    _log_kv("Status", "done  ✅")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
